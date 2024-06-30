import socket
from flask import Flask, render_template, Response, request, jsonify, redirect, url_for, flash, session
import cv2
from pyzbar import pyzbar
from excel_index import query_data_excel, check_active, expired
from project_function import *
from openpyxl import load_workbook
import json
import os
import numpy as np
from send_notification_mail import send_code
import threading
from picamera2 import Picamera2, Preview
from client_pub import publish_pass
from test import get_ipv4_address
def ipadd():
    hostname=socket.gethostname()
    IPAddr=socket.gethostbyname(hostname)
    return IPAddr
path3 = r'/home/pi/DATH/pile.xlsx'
path2 = r'/home/pi/DATH/main_data.xlsx'
path1 = r'/home/pi/DATH/form_register.xlsx'
app = Flask(__name__)
app.secret_key = 'ab156cdth'


picam = Picamera2()
config = picam.create_preview_configuration(main={"size": (1280, 960)})
picam.configure(config)

process_this_frame = True
text= '' #Bien tam de lay du lieu khi quet ma QR
data=[]# Bien tam de lay du lieu tu text sau khi xu ly
data1= []
pile_list =[]
final_list_data = [None] * 13
room_value_send = time_value_send = "none"
room_value = time_value = ""
turn_back = 0#Trả về trang quét mã QR khi admin xác nhận thành công
count = 0 #Trả về trang quét mã QR khi bị admin từ chối 3 lần

training_list = []
folder_name= ''
img_counter = 0
time_counter = 1
face_locations = ''
xac_nhan = '' # Xac nhan chan dung
xac_nhan_khuon_mat = 0 # Nhan dang duoc khuon mat
email= ""
phone = ""
custom_request = ""
piled_request = ""
birth = ""
name = ""
cccd = ""
gender = ""
custom_birth= ""
custom_gender= ""
custom_room_id = ""
custom_check_in = ""
custom_check_out = ""
custom_cccd = ""
custom_name = ""
custom_phone = ""
custom_email = ""
custom_leng_stay = ""
custom_pile_money = 0
#-------------Password----------------
room1_pass = "123456"
room2_pass = "123456"
random_id= '7s76fsg997df7s88f'
randomAdminSerID = '873859178256345'
#Trang chính
@app.route('/', methods=['POST', 'GET'])
def main_page():
    global text
    text = ""
    return render_template("main_page.html")
#Trang đăng ký thông tin------------------------------------------------------------------------------------------
@app.route('/register', methods=['POST', 'GET'])
def register():
    return render_template("register_information.html")
@app.route('/register/choose_room', methods=['POST', 'GET'])
def choose_room():
    global path1, phone, email, birth, name, cccd, gender
    register_list = []
    if request.method == "POST":
        phone = request.form.get("phone")
        email = request.form.get("email")
        birth = request.form.get("birth")
        name = request.form.get("name")
        cccd = request.form.get("cccd")
        gender = request.form.get("gender")
        register_list.append(cccd)
        register_list.append(name)
        register_list.append(birth)
        register_list.append(gender)
        register_list.append(email)
        register_list.append(phone)
        query_data_excel(register_list, path1)
    return render_template("choose_room.html")
#Trang chọn thời gian và tính tiền--------------------------------------------------------------------------------
@app.route('/register/choose_room/<room>', methods=['POST', 'GET'])
def get_room(room):
    time_date = ""
    if room == 'room1':
        if request.method == "POST":
            time_date = request.form.get("date")
        return render_template('room1_registration.html', data = time_date)
    if room == 'room2':
        time_date = request.form.get("date")
        return render_template('room2_registration.html', data = time_date)
@app.route('/register/choose_room/<room>/<calendar>', methods=['POST', 'GET'])
def calendar(room, calendar):
    if room == 'room1' and calendar == "calendar":
        return render_template('calendar.html')
    elif room == 'room2' and calendar == "calendar2":
        return render_template('calendar2.html')
@app.route('/register/choose_room/<room>/pile', methods=['POST', 'GET'])
def pile(room):
    global path1, phone, email, birth, name, cccd, gender
    start_day = ""
    end_day = ""
    leng_stay = ""
    room_id = ""
    total_money = 0
    if room == 'room1':
        if request.method == "POST":
            start_day = request.form.get("start_day")
            leng_stay = request.form.get("length_stay")
            day, month, year = format_date(start_day)
            start_day = day + " " + month + " " + year
            end_day = calendar_operator(day, month, year, leng_stay)
            room_id = request.form.get("room_id")
            total_money = 200000 * int(leng_stay)
            total_money = "{:,}".format(total_money)
    elif room == 'room2':
        if request.method == "POST":
            start_day = request.form.get("start_day")
            leng_stay = request.form.get("length_stay")
            day, month, year = format_date(start_day)
            start_day = day + " " + month + " " + year
            end_day = calendar_operator(day, month, year, leng_stay)
            room_id = request.form.get("room_id")
            total_money = 200000 * int(leng_stay)
            total_money = "{:,}".format(total_money)
    return render_template('pile.html', start_day = start_day, end_day = end_day, leng_stay = leng_stay, total_money = total_money
                           , phone = phone, email = email, birth = birth, name = name, cccd = cccd, gender = gender,
                           room_id = room_id )


#Trang để quét mã QR------------------------------------------------------------------------------------------
@app.route('/qr_detect', methods=['POST', 'GET'])
def qr_detect():
    global data, text, path3
    if request.method == "POST":
        if text != "":
            data = step_list(text)
            print(data)
            wb = load_workbook(path3)
            ws = wb.active
            for i in range(2, ws.max_row + 1):
                row_status = [cell.value for cell in ws[i][0:12]]
                print(row_status)
                if data[0] in row_status and format_and_compare(row_status[1]) == True:
                    print(row_status[1])
                    return jsonify({"status": "piled"})
    return render_template('qr_detect.html')
def gen2():
    """Video streaming generator function."""
    global text, process_this_frame, directory, img_counter, time_counter, directory1, known_face_encodings, xac_nhan_khuon_mat
    #picam.start_preview(Preview.QTGL)
    picam.start()
    while True:
        picam.capture_file("t.jpg")
        img = cv2.imread('t.jpg')
        barcodes = pyzbar.decode(img)
        for barcode in barcodes:
            (x, y, w, h) = barcode.rect
            cv2.rectangle(img, (x, y), (x + w, y + h), (0, 0, 255), 2)
            barcodeData = barcode.data.decode("utf-8")
            text = "{}".format(barcodeData)
            print(text)
            cv2.putText(img, text, (x, y - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + open('t.jpg', 'rb').read() + b'\r\n')
@app.route('/video_feed2')
def video_feed2():
    return Response(gen2(),
                    mimetype='multipart/x-mixed-replace; boundary=frame')

#Trang hiển thị thông tin------------------------------------------------------------------------------------------
@app.route('/display_data', methods=['POST', 'GET'])
def index1():
    global data, text, xac_nhan, known_face_encodings
    known_face_encodings = []
    xac_nhan= ''
    data = step_list(text)
    if request.method == "POST":
        return jsonify({'cccd': data[0], 'name': data[2], 'born':data[3], 'gender': data[4]})
    return render_template('display_data.html')
@app.route('/customer_piled', methods=['POST', 'GET'])
def index():
    global data, text, path3
    row_status = []
    data = step_list(text)
    wb = load_workbook(path3)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        row_status = [cell.value for cell in ws[i][0:12]]
    return render_template('display_data_piled.html', start_day=row_status[1], end_day=row_status[2], leng_stay=row_status[3],
                           total_money=row_status[10],
                            phone=row_status[9], email=row_status[8], birth=row_status[6], name=row_status[5], cccd=row_status[4], gender=row_status[7],
                           room_id=row_status[11])
#Trang cảm ơn------------------------------------------------------------------------------------------
@app.route('/camon', methods=['POST', 'GET'])
def camon():
    global data, data1, email, text
    text = ""
    data1 = data.copy()
    if request.method == "POST":
        std = request.form.get("std")
        email = request.form.get("email")
        data1.append(std)
        data1.append(email)
    return render_template('camon.html')
@app.route('/final_payment', methods=['POST', 'GET'])
def final_payment():
    global data, data1, email, text, final_list_data
    final_list_data = [None] * 11
    text = ""
    data1 = data.copy()
    leng_stay = 0
    if request.method == "POST":
        leng_stay = request.form.get("leng_stay")
        money = int(leng_stay) * 200000
        total_money = money * 2
        money = "{:,}".format(money)
        final_list_data[0] = request.form.get("start_day")
        final_list_data[1] = request.form.get("end_day")
        final_list_data[2] = request.form.get("leng_stay")
        final_list_data[3] = request.form.get("cccd")
        final_list_data[4] = request.form.get("fname")
        final_list_data[5] = request.form.get("birth")
        final_list_data[6] = request.form.get("gender")
        final_list_data[7] = request.form.get("email")
        final_list_data[8] = request.form.get("std")
        final_list_data[9] = total_money
        final_list_data[10] = request.form.get("room_id")
    return render_template('final_payment.html', leng_stay= leng_stay , money=money)

#---------------------------------------Admin------------------------------------------
@app.route('/admin_login', methods=['POST', 'GET'])
def admin_login():
    global randomAdminSerID
    error = None
    if request.method == 'POST':
        if request.form['user'] == 'Admin' and \
                request.form['password'] == '777777':
            randomAdminSerID = pin_random(20)
            return redirect(url_for('admin', randomAdminSerID = randomAdminSerID))
        else:
            error = 'Invalid User or password. Please try again!'
    return render_template('admin_sign_in.html', error=error)
@app.route('/admin/<randomAdminSerID>', methods=['POST', 'GET'])
#Trang admin----------------------------------------------------------------
def admin(randomAdminSerID):
    global room1_pass, room2_pass, pile_list, final_list_data, piled_request, time_value_send, room_value_send, data, data1, room_value, path2, path3,custom_check_out, custom_room_id, custom_request, time_value, turn_back, count, email, path2, custom_name, custom_cccd, custom_birth, custom_gender, custom_email, custom_phone, custom_check_in, custom_leng_stay, custom_pile_money
    room1_cccd = room1_name = room1_id = room1_check_in = room1_check_out = ""
    room2_cccd = room2_name = room2_id = room2_check_in = room2_check_out = ""
    wb = load_workbook(path2)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        row_status = [cell.value for cell in ws[i][0:14]]
        if row_status[-2] == 'active':
            if row_status[-3] == 'room1':
                room1_cccd = row_status[4]
                room1_name = row_status[5]
                room1_id = row_status[-3]
                room1_check_in = row_status[1]
                room1_check_out = row_status[2]
                room1_pass = row_status[-1]
            if row_status[-3] == 'room2':
                room2_cccd = row_status[4]
                room2_name = row_status[5]
                room2_id = row_status[-3]
                room2_check_in = row_status[1]
                room2_check_out = row_status[2]
                room2_pass = row_status[-1]

    wb.save(path2)
    t2 = threading.Thread(target=check_active, args=(path2,))
    t2.start()
    t2.join()
    if request.method == "POST":
        verify = json.loads(request.data).get('verify')
        final_verify = json.loads(request.data).get('final_verify')
        if verify == "COMPLETE":
            turn_back = 1
            pile_list_copy = pile_list.copy()
            query_data_excel(pile_list_copy, path3)
            custom_request = "NO"
        if final_verify == "COMPLETE":
            turn_back = 1
            password = pin_random(6)
            final_list_data_copy = final_list_data.copy()
            final_list_data_copy.append('active')
            final_list_data_copy.append(password)
            print(final_list_data_copy)
            query_data_excel(final_list_data_copy, path2)
            send_code(final_list_data[7], password)
            #publish_pass(final_list_data[10], password)
            piled_reques0t = "NO"
        return jsonify({ 'room1_cccd': room1_cccd, 'room1_name' : room1_name, 'room1_id': room1_id, 'room1_check_in': room1_check_in, 'room1_check_out' : room1_check_out,
                         'room2_cccd': room2_cccd, 'room2_name': room2_name, 'room2_id': room2_id,
                         'room2_check_in': room2_check_in, 'room2_check_out': room2_check_out,
                        'piled_request' : piled_request, "final_cccd" : final_list_data[3], "final_room_id": final_list_data[10], "remain_payment" : "{:,}".format(int(final_list_data[9])/2) if final_list_data[9] != None else None,
                         "final_check_in": final_list_data[0],
                        'pile_cccd': custom_cccd,'custom_request': custom_request, "pile_phone_number": custom_phone,
                        'room': room_value, 'time':time_value, "pile_room" : custom_room_id, "pile_start_day": custom_check_in,
                        "pile_end_day": custom_check_out, "pile_leng_stay": custom_leng_stay, "pile_total_money": str(custom_pile_money)})

    return render_template('admin.html')
@app.route('/phong_cho', methods=['POST', 'GET'])
def xac_nhan():
    global pile_list, piled_request, time_value_send,room_value_send ,turn_back, count,custom_request,custom_name,custom_room_id, custom_cccd, custom_birth, custom_gender, custom_email, custom_phone, custom_check_in, custom_check_out, custom_leng_stay, custom_pile_money
    if request.method == 'POST':
        room_value_send = request.form.get("room")
        time_value_send = request.form.get("time")
        custom_request = request.form.get("request")
        custom_name = request.form.get("fname")
        custom_cccd = request.form.get("cccd")
        custom_birth = request.form.get("birth")
        custom_gender = request.form.get("gender")
        custom_email = request.form.get("email")
        custom_phone = request.form.get("phone")
        custom_room_id = request.form.get("room_id")
        custom_check_in = request.form.get("start_day")
        custom_check_out = request.form.get("end_day")
        custom_leng_stay = request.form.get("leng_stay")
        custom_pile_money = request.form.get("total_money")
        piled_request = request.form.get("piled_request")
        if custom_name != None:
            pile_list = list_insert(custom_check_in,custom_check_out, custom_leng_stay,custom_cccd, custom_name, custom_birth, custom_gender, custom_email, custom_phone, custom_pile_money, custom_room_id)
            print(pile_list)
        if turn_back == 1 or count == 4:
            turn_back = 0
            count = 0
            return jsonify({"status": "turnback"})
    return render_template('phong_cho.html')
#------------------------Check-out-----------------
@app.route('/check-out', methods=['POST', 'GET'])
def check_out():
    global room1_pass, room2_pass, random_id, path2
    error = None
    if request.method == 'POST':
        session["room_id"] = request.form['room_id']
        if request.form['room_id'] == 'room1' and \
                request.form['password'] == room1_pass:
            random_id = pin_random(10)
            return redirect(url_for('final_check', room_id= 'room1', random_pin= random_id))
        elif request.form['room_id'] == 'room2' and \
                request.form['password'] == room2_pass:
            random_id = pin_random(10)
            return redirect(url_for('final_check', room_id= 'room2', random_pin= random_id))
        else:
            error = 'Invalid room-id or password. Please try again!'
    return render_template('check_out_sign_in.html', error=error)

@app.route(f'/check-out/<room_id>/<random_pin>', methods=['POST', 'GET'])
def final_check(room_id, random_pin):
    global random_id
    if room_id == "room1" and random_pin == random_id:
        if request.form.get('room') == "Submit":
            expired(path2, "room1")
            return redirect('/')
        return render_template("room1_check_out.html")
    if room_id == "room2" and random_pin == random_id:
        if request.form.get('room') == "Submit":
            expired(path2, "room2")
            return redirect('/')
        return render_template("room2_check_out.html")


if __name__ == '__main__':
    app.run(host=get_ipv4_address(), debug=False)#f'{ipadd()}'
