from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from project_function import step_list
import time
from client_pub import publish_pass                                                            
def query_data_excel(list, path):
    list.insert(0, datetime.now())
    wb = load_workbook(path)
    ws = wb.active
    #Lay du lieu trong database(excel)
    ws.append(list)
    wb.save(path)
    #------------------------------------
    '''elif(check_exist1(data, lst) == None):
        for i in range(len(lst1)):
            if lst[i] == data:
                lst2 = final_list(data, lst1[i][0], lst1[i][1])
                if lst2 != None:
                    for col in range(0,2):
                        char = get_column_letter(col+1)
                        ws[char + str(i+1)] = lst2[col]
                    wb.save('C:/Users/Huy/Documents/Doan.xlsx')'''
def check_active(path):
    calendar = {
        "Jan": 1,
        "Feb": 2,
        "Mar": 3,
        "Apr": 4,
        "May": 5,
        "Jun": 6,
        "Jul": 7,
        "Aug": 8,
        "Sep": 9,
        "Oct": 10,
        "Nov": 11,
        "Dec": 12
    }
    wb = load_workbook(path)
    ws = wb.active
    row_status = []
    for i in range(2, ws.max_row + 1):
        row_status = [cell.value for cell in ws[i][0:14]]
        if row_status[-3] == 'room1' and row_status[-2] == 'active':
            publish_pass(row_status[-3], row_status[-1])
        if row_status[-3] == 'room2' and row_status[-2] == 'active':
            publish_pass(row_status[-3], row_status[-1])
        if row_status[-1] == 'active':
            list_date = str(row_status[2]).split()
            index = list(calendar).index(list_date[1])
            month = list(calendar.values())[index]
            check_out_time = datetime(int(list_date[2]), int(month), int(list_date[0]),12)
            if datetime.now() > check_out_time:
                print(row_status)
                ws["M" + str(i)] = "expired"
                ws["N" + str(i)] = "None"
            else:
                print(None)
    wb.save(path)
def delete_row(path):
    wb = load_workbook(path)
    ws = wb.active
    ws.delete_rows(2,1000)
    wb.save(path)

path3 = r'/home/pi/DATH/pile.xlsx'
path2 = r'/home/pi/DATH/main_data.xlsx'
path1 = r'/home/pi/DATH/form_register.xlsx'
#delete_row(path2)
check_active(path2)

def expired(path, room_id):
    wb = load_workbook(path)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        row_status = [cell.value for cell in ws[i][0:13]]
        if row_status[-1] == 'active' and row_status[-2] == room_id:
            ws["M" + str(i)] = "expired"
            ws["N" + str(i)] = "None"
            ws["O" + str(i)] = datetime.now()
            publish_pass(room_id, "None")
    wb.save(path)
